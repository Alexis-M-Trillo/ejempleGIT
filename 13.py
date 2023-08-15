from PIL import Image
import openpyxl

# Ruta de la imagen
ruta_imagen = "C:/Users/a/Documents/Practica/pingui01.png"

# Tamaño máximo deseado para la imagen
tamaño_maximo = (200, 200)  # Especifica el tamaño máximo deseado (ancho, alto)

# Abrir la imagen y redimensionarla
imagen = Image.open(ruta_imagen)
imagen.thumbnail(tamaño_maximo)

# Obtener los píxeles de la imagen
pixeles = imagen.load()

# Crear un nuevo archivo de Excel
libro_excel = openpyxl.Workbook()
hoja_excel = libro_excel.active

# Recorrer cada píxel de la imagen
for fila in range(imagen.height):
    for columna in range(imagen.width):
        # Obtener el color del píxel
        color = pixeles[columna, fila]

        # Obtener los valores RGB del color
        r, g, b = color[0], color[1], color[2]

        # Crear un estilo de celda con el color RGB
        estilo_celda = openpyxl.styles.PatternFill(start_color=f"00{r:02X}{g:02X}{b:02X}",
                                                   end_color=f"00{r:02X}{g:02X}{b:02X}",
                                                   fill_type="solid")

        # Guardar los valores en la tabla de Excel y aplicar el estilo de celda
        celda = hoja_excel.cell(row=fila+1, column=columna+1)
        celda.fill = estilo_celda

# Ajustar el ancho de las columnas y el alto de las filas para que coincidan con el tamaño de la imagen
for i in range(1, imagen.width + 1):
    hoja_excel.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 1
for i in range(1, imagen.height + 1):
    hoja_excel.row_dimensions[i].height = 1

# Eliminar el texto de las celdas
for fila in hoja_excel.iter_rows():
    for celda in fila:
        celda.value = None

# Guardar el archivo de Excel
ruta_excel = "C:/Users/a/Documents/Practica/prueba.xlsx"
libro_excel.save(ruta_excel)

print("¡Archivo de Excel creado exitosamente!")
