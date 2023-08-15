from PIL import Image
import openpyxl

# Ruta de la imagen
ruta_imagen = "C:/Users/a/Documents/Practica/112.png"

# Abrir la imagen
imagen = Image.open(ruta_imagen)

# Reducir la imagen a 32x32 píxeles
imagen_reducida = imagen.resize((32, 32))

# Obtener los píxeles de la imagen reducida
pixeles = imagen_reducida.load()

# Crear un nuevo archivo de Excel
libro_excel = openpyxl.Workbook()
hoja_excel = libro_excel.active

# Recorrer cada píxel de la imagen reducida
for fila in range(imagen_reducida.height):
    for columna in range(imagen_reducida.width):
        # Obtener el color del píxel
        color = pixeles[columna, fila]

        # Obtener los valores RGB del color
        r, g, b = color[0], color[1], color[2]

        # Crear un estilo de celda con el color RGB
        estilo_celda = openpyxl.styles.PatternFill(start_color=f"00{r:02X}{g:02X}{b:02X}",
                                                   end_color=f"00{r:02X}{g:02X}{b:02X}",
                                                   fill_type="solid")

        # Guardar los valores en la tabla de Excel y aplicar el estilo de celda
        celda = hoja_excel.cell(row=fila+1, column=columna+1)
        celda.value = f"RGB({r}, {g}, {b})"
        celda.fill = estilo_celda

# Ajustar el ancho de las columnas para que se ajusten a la imagen
for columna in hoja_excel.columns:
    hoja_excel.column_dimensions[columna[0].column_letter].width = 1

# Guardar el archivo de Excel
ruta_excel = "C:/Users/a/Documents/Practica/prueba.xlsx"
libro_excel.save(ruta_excel)

print("¡Archivo de Excel creado exitosamente!")
